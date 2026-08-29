from __future__ import annotations

import csv
import io
import json
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Literal

from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.exceptions import ApplicationError
from app.modules.batch_jobs.schemas import ParsedProductImportRow

SCHEMA_VERSION = "product-import-v1"
MAX_ROWS = 10_000
MAX_COLUMNS = 14
HEADERS = (
    "product_key",
    "product_name",
    "subtitle",
    "category_id",
    "brand_id",
    "merchant_sku_code",
    "sku_name",
    "spec_values_json",
    "sale_price_amount",
    "market_price_amount",
    "currency",
    "weight_grams",
    "barcode",
    "initial_stock",
)


@dataclass(frozen=True)
class RowProblem:
    row_number: int
    code: str
    message: str
    input_hash: bytes


@dataclass(frozen=True)
class ProductImportParseResult:
    rows: tuple[ParsedProductImportRow, ...]
    problems: tuple[RowProblem, ...]


def parse_product_import(payload: bytes, content_type: str) -> ProductImportParseResult:
    if content_type == "text/csv":
        raw_rows = _csv_rows(payload)
    elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        raw_rows = _xlsx_rows(payload)
    else:
        raise _invalid_file("IMPORT_FILE_TYPE_INVALID", "导入文件必须为 CSV 或 XLSX。")
    return _validate_rows(raw_rows)


def _csv_rows(payload: bytes) -> list[tuple[int, dict[str, object]]]:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise _invalid_file("IMPORT_CSV_ENCODING_INVALID", "CSV 必须使用 UTF-8 编码。") from exc
    csv.field_size_limit(10_000)
    reader = csv.reader(io.StringIO(text, newline=""), strict=True)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise _invalid_file("IMPORT_FILE_EMPTY", "导入文件不能为空。") from exc
    except csv.Error as exc:
        raise _invalid_file("IMPORT_CSV_INVALID", "CSV 文件结构无效。") from exc
    _validate_headers(header)
    rows: list[tuple[int, dict[str, object]]] = []
    try:
        for row_number, values in enumerate(reader, start=2):
            if row_number > MAX_ROWS + 1:
                raise _invalid_file("IMPORT_ROW_LIMIT_EXCEEDED", "单次导入最多允许 10000 行。")
            if not any(value.strip() for value in values):
                continue
            if len(values) != len(HEADERS):
                rows.append((row_number, {"__row_error__": "列数与模板不一致。"}))
                continue
            rows.append((row_number, dict(zip(HEADERS, values, strict=True))))
    except csv.Error as exc:
        raise _invalid_file("IMPORT_CSV_INVALID", "CSV 文件结构无效。") from exc
    return rows


def _xlsx_rows(payload: bytes) -> list[tuple[int, dict[str, object]]]:
    _preflight_xlsx(payload)
    try:
        workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise _invalid_file("IMPORT_XLSX_INVALID", "XLSX 文件结构无效。") from exc
    try:
        if len(workbook.worksheets) != 1:
            raise _invalid_file("IMPORT_XLSX_SHEET_COUNT_INVALID", "导入模板只能包含一个工作表。")
        worksheet = workbook.worksheets[0]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = ["" if value is None else str(value).strip() for value in next(iterator)]
        except StopIteration as exc:
            raise _invalid_file("IMPORT_FILE_EMPTY", "导入文件不能为空。") from exc
        _validate_headers(header)
        rows: list[tuple[int, dict[str, object]]] = []
        for row_number, values in enumerate(iterator, start=2):
            if row_number > MAX_ROWS + 1:
                raise _invalid_file("IMPORT_ROW_LIMIT_EXCEEDED", "单次导入最多允许 10000 行。")
            normalized = tuple(values[:MAX_COLUMNS])
            if len(values) > MAX_COLUMNS and any(
                value is not None for value in values[MAX_COLUMNS:]
            ):
                rows.append((row_number, {"__row_error__": "列数与模板不一致。"}))
                continue
            if not any(value is not None and str(value).strip() for value in normalized):
                continue
            padded = normalized + (None,) * (len(HEADERS) - len(normalized))
            rows.append((row_number, dict(zip(HEADERS, padded, strict=True))))
        return rows
    finally:
        workbook.close()


def _preflight_xlsx(payload: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            if len(members) > 500:
                raise _invalid_file("IMPORT_XLSX_ARCHIVE_LIMIT", "XLSX 内部文件数量超过安全上限。")
            total_size = 0
            for member in members:
                normalized = member.filename.casefold()
                if "vbaproject.bin" in normalized or normalized.startswith("xl/externallinks/"):
                    raise _invalid_file(
                        "IMPORT_XLSX_ACTIVE_CONTENT_FORBIDDEN",
                        "XLSX 不允许包含宏或外部链接。",
                    )
                total_size += member.file_size
                if member.file_size > 100 * 1024 * 1024 or total_size > 200 * 1024 * 1024:
                    raise _invalid_file(
                        "IMPORT_XLSX_ARCHIVE_LIMIT", "XLSX 解压后大小超过安全上限。"
                    )
                if member.compress_size and member.file_size / member.compress_size > 500:
                    raise _invalid_file("IMPORT_XLSX_ARCHIVE_LIMIT", "XLSX 压缩比超过安全上限。")
    except zipfile.BadZipFile as exc:
        raise _invalid_file("IMPORT_XLSX_INVALID", "XLSX 文件结构无效。") from exc


def _validate_headers(header: Iterable[object]) -> None:
    normalized = tuple(str(value).strip() for value in header)
    if normalized != HEADERS:
        raise _invalid_file(
            "IMPORT_SCHEMA_MISMATCH",
            f"模板列必须严格匹配 {SCHEMA_VERSION}，请下载最新模板。",
        )


def _validate_rows(
    raw_rows: list[tuple[int, dict[str, object]]],
) -> ProductImportParseResult:
    if not raw_rows:
        raise _invalid_file("IMPORT_NO_DATA_ROWS", "导入文件没有数据行。")
    rows: list[ParsedProductImportRow] = []
    problems: list[RowProblem] = []
    seen_sku_codes: dict[str, int] = {}
    product_signatures: dict[str, tuple[str, str, str | None, str | None]] = {}
    product_spec_signatures: dict[str, dict[str, int]] = {}
    invalid_product_keys: set[str] = set()
    row_by_number: dict[int, ParsedProductImportRow] = {}
    for row_number, raw in raw_rows:
        digest = _row_hash(raw)
        if "__row_error__" in raw:
            problems.append(
                RowProblem(
                    row_number, "IMPORT_ROW_SHAPE_INVALID", str(raw["__row_error__"]), digest
                )
            )
            continue
        try:
            spec_values = json.loads(_text(raw.get("spec_values_json"), required=True))
            if not isinstance(spec_values, list) or not all(
                isinstance(item, dict) for item in spec_values
            ):
                raise ValueError("spec_values_json must be an array of objects")
            row = ParsedProductImportRow(
                row_number=row_number,
                product_key=_text(raw.get("product_key"), required=True, max_length=64),
                product_name=_text(raw.get("product_name"), required=True, max_length=255),
                subtitle=_optional_text(raw.get("subtitle"), 500),
                category_id=_text(raw.get("category_id"), required=True, max_length=40),
                brand_id=_optional_text(raw.get("brand_id"), 40),
                merchant_sku_code=_text(raw.get("merchant_sku_code"), required=True, max_length=64),
                sku_name=_text(raw.get("sku_name"), required=True, max_length=255),
                spec_values=[
                    {
                        "name": str(item.get("name", "")).strip(),
                        "value": str(item.get("value", "")).strip(),
                    }
                    for item in spec_values
                ],
                sale_price_amount=_integer(raw.get("sale_price_amount"), minimum=0),
                market_price_amount=_integer(raw.get("market_price_amount"), minimum=0),
                currency=_currency(raw.get("currency")),
                weight_grams=_optional_integer(raw.get("weight_grams"), minimum=0),
                barcode=_optional_text(raw.get("barcode"), 64),
                initial_stock=_integer(raw.get("initial_stock"), minimum=0, maximum=1_000_000_000),
            )
        except (ValueError, TypeError, json.JSONDecodeError, ValidationError) as exc:
            problems.append(
                RowProblem(row_number, "IMPORT_ROW_INVALID", _safe_validation_message(exc), digest)
            )
            continue
        duplicate_row = seen_sku_codes.get(row.merchant_sku_code.casefold())
        if duplicate_row is not None:
            problems.append(
                RowProblem(
                    row_number,
                    "IMPORT_SKU_CODE_DUPLICATE",
                    f"SKU 编码与第 {duplicate_row} 行重复。",
                    digest,
                )
            )
            continue
        seen_sku_codes[row.merchant_sku_code.casefold()] = row_number
        signature = (row.product_name, row.category_id, row.brand_id, row.subtitle)
        previous = product_signatures.get(row.product_key)
        if previous is not None and previous != signature:
            problems.append(
                RowProblem(
                    row_number,
                    "IMPORT_PRODUCT_FIELDS_CONFLICT",
                    "同一 product_key 的商品字段不一致。",
                    digest,
                )
            )
            invalid_product_keys.add(row.product_key)
            continue
        product_signatures[row.product_key] = signature
        spec_signature = json.dumps(
            sorted(
                (
                    item["name"].strip().casefold(),
                    item["value"].strip().casefold(),
                )
                for item in row.spec_values
            ),
            ensure_ascii=False,
            separators=(",", ":"),
        )
        previous_spec_row = product_spec_signatures.setdefault(row.product_key, {}).get(
            spec_signature
        )
        if previous_spec_row is not None:
            problems.append(
                RowProblem(
                    row_number,
                    "IMPORT_SPEC_COMBINATION_DUPLICATE",
                    f"规格组合与第 {previous_spec_row} 行重复。",
                    digest,
                )
            )
            invalid_product_keys.add(row.product_key)
            continue
        product_spec_signatures[row.product_key][spec_signature] = row_number
        row_by_number[row_number] = row
        rows.append(row)
    invalid_products = invalid_product_keys | {
        row_by_number[problem.row_number].product_key
        for problem in problems
        if problem.row_number in row_by_number
    }
    if invalid_products:
        kept: list[ParsedProductImportRow] = []
        existing_problem_rows = {problem.row_number for problem in problems}
        for row in rows:
            if row.product_key in invalid_products:
                if row.row_number not in existing_problem_rows:
                    problems.append(
                        RowProblem(
                            row.row_number,
                            "IMPORT_PRODUCT_GROUP_INVALID",
                            "同一商品组存在错误，本行不会执行。",
                            _row_hash(row.model_dump(mode="json")),
                        )
                    )
            else:
                kept.append(row)
        rows = kept
    return ProductImportParseResult(
        tuple(rows), tuple(sorted(problems, key=lambda item: item.row_number))
    )


def _text(value: object, *, required: bool, max_length: int = 10_000) -> str:
    text = "" if value is None else str(value).strip()
    if required and not text:
        raise ValueError("required value is empty")
    if len(text) > max_length:
        raise ValueError("value exceeds maximum length")
    return text


def _currency(value: object) -> Literal["CNY"]:
    currency = _text(value, required=True, max_length=3)
    if currency != "CNY":
        raise ValueError("currency must be CNY")
    return "CNY"


def _optional_text(value: object, max_length: int) -> str | None:
    text = _text(value, required=False, max_length=max_length)
    return text or None


def _integer(value: object, *, minimum: int, maximum: int = 10**15) -> int:
    if isinstance(value, bool):
        raise ValueError("boolean is not an integer")
    if isinstance(value, float) and not value.is_integer():
        raise ValueError("integer field contains a decimal")
    if isinstance(value, float):
        parsed = int(value)
        if parsed < minimum or parsed > maximum:
            raise ValueError("integer field is outside the allowed range")
        return parsed
    text = _text(value, required=True)
    if text.startswith("+"):
        text = text[1:]
    if not text.isdigit():
        raise ValueError("integer field is invalid")
    parsed = int(text)
    if parsed < minimum or parsed > maximum:
        raise ValueError("integer field is outside the allowed range")
    return parsed


def _optional_integer(value: object, *, minimum: int) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    return _integer(value, minimum=minimum)


def _row_hash(raw: object) -> bytes:
    import hashlib

    return hashlib.sha256(
        json.dumps(
            raw, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
        ).encode()
    ).digest()


def _safe_validation_message(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        first = exc.errors(include_url=False)[0]
        field = ".".join(str(item) for item in first.get("loc", ()))
        return f"字段 {field or 'row'} 不符合模板约束。"
    return "字段值不符合模板约束。"


def _invalid_file(code: str, detail: str) -> ApplicationError:
    return ApplicationError(
        status=422,
        code=code,
        title="Invalid import file",
        detail=detail,
    )
